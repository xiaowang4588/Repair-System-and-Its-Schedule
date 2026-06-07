"""
报修管理模块

负责：
  1. 报修记录的增删改查（Peewee ORM / SQLite 存储）
  2. 智能填充（根据教室查询当前课程/教师）
  3. 就近空教室推荐（同楼栋同层优先）
  4. Excel 导出（与原始格式一致）
  5. Excel 导入（含数据验证和错误处理）
  6. 报修统计
"""
import json
import os
import re
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import db, Repair, ModificationLog
from utils.stats_helper import filter_by_range, count_dict, trend_data, avg_process_days, extract_building

logger = logging.getLogger(__name__)

# 项目根目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _cleanup_removed_images(old_images_json: str, new_images_json: str):
    """编辑时清理被移除的图片文件"""
    try:
        old_list = json.loads(old_images_json) if old_images_json else []
        new_list = json.loads(new_images_json) if new_images_json else []
        for img in old_list:
            if img not in new_list:
                _delete_upload_file(img)
    except Exception:
        pass


def _delete_upload_file(path: str):
    """删除 uploads 目录下的单个文件"""
    if not path or not path.startswith('/uploads/'):
        return
    full_path = os.path.join(os.path.dirname(BASE_DIR), path.lstrip('/'))
    try:
        if os.path.exists(full_path):
            os.remove(full_path)
            logger.info(f"已清理孤立图片: {full_path}")
    except Exception as e:
        logger.warning(f"删除图片失败 {full_path}: {e}")


# ============================================================
# SQL 查询辅助
# ============================================================

def _sql_date_filter(range_type: str):
    """返回 Peewee where 条件，按时间范围过滤"""
    now = datetime.now()
    if range_type == 'week':
        week_start = (now - timedelta(days=now.weekday())).strftime('%Y-%m-%d')
        return Repair.report_time >= week_start
    elif range_type == 'month':
        month_start = now.strftime('%Y-%m-01')
        return Repair.report_time >= month_start
    return True  # semester: 无日期过滤


def _to_light_dicts(query) -> list:
    """将 Peewee 查询结果转为轻量字典列表（无修改日志 N+1）"""
    return [r.to_dict_light() for r in query]


# ============================================================
# 报修记录 CRUD
# ============================================================

def get_repair_list(status: str = 'all',
                    semester: str = '',
                    start_date: str = '',
                    end_date: str = '',
                    keyword: str = '',
                    classroom: str = '',
                    week_number: str = '',
                    fault_type: str = '',
                    reporter_name: str = '',
                    reporter_college: str = '',
                    is_external_teacher: str = '',
                    handler_name: str = '',
                    fault_cause: str = '',
                    page: int = 1,
                    page_size: int = 50) -> dict:
    """
    获取报修列表（支持多条件组合筛选和分页）

    :return: {records: list, total: int, page: int, page_size: int, total_pages: int}
    """
    query = Repair.select()

    # 按学期筛选
    if semester and semester != 'all':
        query = query.where(Repair.semester == semester)

    # 按状态筛选
    if status and status != 'all':
        query = query.where(Repair.status == status)

    # 按时间范围筛选
    if start_date:
        query = query.where(Repair.report_time >= start_date)
    if end_date:
        query = query.where(Repair.report_time <= end_date + ' 23:59:59')

    # 按教室筛选
    if classroom:
        query = query.where(Repair.classroom.contains(classroom))

    # 按周次筛选
    if week_number:
        try:
            query = query.where(Repair.week_number == int(week_number))
        except ValueError:
            pass

    # 按报修类型筛选
    if fault_type and fault_type != 'all':
        query = query.where(Repair.fault_type == fault_type)

    # 按报修人筛选
    if reporter_name:
        query = query.where(Repair.reporter_name.contains(reporter_name))

    # 按报修人学院筛选
    if reporter_college and reporter_college != 'all':
        query = query.where(Repair.reporter_college.contains(reporter_college))

    # 按外聘教师筛选
    if is_external_teacher and is_external_teacher != 'all':
        query = query.where(Repair.is_external_teacher == (is_external_teacher == 'true'))

    # 按处理人筛选
    if handler_name:
        query = query.where(Repair.handler_name.contains(handler_name))

    # 按故障原因筛选
    if fault_cause:
        query = query.where(Repair.fault_cause.contains(fault_cause))

    # 按关键词搜索（通用搜索）
    if keyword:
        query = query.where(
            (Repair.classroom.contains(keyword)) |
            (Repair.reporter_name.contains(keyword)) |
            (Repair.fault_cause.contains(keyword)) |
            (Repair.solution.contains(keyword)) |
            (Repair.handler_name.contains(keyword))
        )

    # 分页
    total = query.count()
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))

    records = query.order_by(Repair.report_time.desc()).paginate(page, page_size)

    return {
        'records': [r.to_dict() for r in records],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }


def get_filter_options() -> dict:
    """获取所有筛选选项（用于前端下拉框）"""
    # 学期列表
    semesters = sorted(
        set(r.semester for r in Repair.select(Repair.semester).distinct() if r.semester),
        reverse=True
    )

    # 报修类型列表
    fault_types = sorted(
        set(r.fault_type for r in Repair.select(Repair.fault_type).distinct() if r.fault_type)
    )

    # 学院列表
    colleges = sorted(
        set(r.reporter_college for r in Repair.select(Repair.reporter_college).distinct() if r.reporter_college)
    )

    # 处理状态列表
    statuses = sorted(
        set(r.status for r in Repair.select(Repair.status).distinct() if r.status)
    )

    # 教室列表（去重）
    classrooms = sorted(
        set(r.classroom for r in Repair.select(Repair.classroom).distinct() if r.classroom)
    )

    # 处理人列表
    handlers = sorted(
        set(r.handler_name for r in Repair.select(Repair.handler_name).distinct() if r.handler_name)
    )

    # 报修人列表
    reporters = sorted(
        set(r.reporter_name for r in Repair.select(Repair.reporter_name).distinct() if r.reporter_name)
    )

    # 周次列表
    weeks = sorted(
        set(r.week_number for r in Repair.select(Repair.week_number).distinct() if r.week_number and r.week_number > 0)
    )

    return {
        'semesters': semesters,
        'fault_types': fault_types,
        'colleges': colleges,
        'statuses': statuses,
        'classrooms': classrooms,
        'handlers': handlers,
        'reporters': reporters,
        'weeks': weeks,
    }


def batch_update_status(record_ids: list, new_status: str) -> dict:
    """批量更新处理状态（带事务保护）"""
    from models import get_db
    db = get_db()
    try:
        with db.atomic():
            updated = Repair.update(
                status=new_status,
                updated_at=datetime.now().strftime('%Y-%m-%d')
            ).where(Repair.id.in_([int(rid) for rid in record_ids])).execute()

            if updated > 0:
                logger.info(f"批量更新状态: {updated} 条记录 -> {new_status}")

            return {'updated_count': updated}
    except Exception as e:
        logger.error(f"批量更新状态失败: {e}")
        return {'error': str(e), 'updated_count': 0}


def batch_update_handler(record_ids: list, handler_name: str) -> dict:
    """批量分配处理人（带事务保护）"""
    from models import get_db
    db = get_db()
    try:
        with db.atomic():
            updated = Repair.update(
                handler_name=handler_name,
                updated_at=datetime.now().strftime('%Y-%m-%d')
            ).where(Repair.id.in_([int(rid) for rid in record_ids])).execute()

            if updated > 0:
                logger.info(f"批量分配处理人: {updated} 条记录 -> {handler_name}")

            return {'updated_count': updated}
    except Exception as e:
        logger.error(f"批量分配处理人失败: {e}")
        return {'error': str(e), 'updated_count': 0}


def batch_delete(record_ids: list) -> dict:
    """批量删除（带事务保护）"""
    from models import get_db
    db = get_db()
    try:
        with db.atomic():
            deleted = Repair.delete().where(
                Repair.id.in_([int(rid) for rid in record_ids])
            ).execute()

            if deleted > 0:
                logger.info(f"批量删除: {deleted} 条记录")

            return {'deleted_count': deleted}
    except Exception as e:
        logger.error(f"批量删除失败: {e}")
        return {'error': str(e), 'deleted_count': 0}


def get_semester_list() -> list:
    """获取所有学期列表（用于筛选）"""
    semesters = sorted(
        set(r.semester for r in Repair.select(Repair.semester).distinct() if r.semester),
        reverse=True
    )
    return semesters


def _get_current_semester() -> str:
    """获取当前学期名称，如 '2025-2026春季'"""
    import services.admin_config as admin_config
    try:
        semester_config = admin_config.get_semester_config()
        start_date = semester_config.get('start_date', '')
        if start_date:
            year = int(start_date[:4])
            month = int(start_date[5:7])
            if month >= 8:  # 8月开始是秋季
                return f"{year}-{year+1}秋季"
            else:  # 2月开始是春季
                return f"{year-1}-{year}春季"
    except Exception:
        pass
    # 默认根据当前时间判断
    now = datetime.now()
    if now.month >= 8:
        return f"{now.year}-{now.year+1}秋季"
    else:
        return f"{now.year-1}-{now.year}春季"


def create_repair(data: dict) -> dict:
    """创建报修记录"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    repair = Repair.create(
        student_id=data.get('student_id', ''),
        student_name=data.get('student_name', ''),
        semester=data.get('semester', _get_current_semester()),
        classroom=data.get('classroom', ''),
        report_time=data.get('report_time', datetime.now().strftime('%Y-%m-%d')),
        week_number=int(data.get('week_number', 0)),
        fault_type=data.get('fault_type', ''),
        reporter_name=data.get('reporter_name', ''),
        reporter_college=data.get('reporter_college', ''),
        is_external_teacher=bool(data.get('is_external_teacher', False)),
        report_method=data.get('report_method', ''),
        handler_name=data.get('handler_name', ''),
        is_device_replaced=bool(data.get('is_device_replaced', False)),
        device_replace_note=data.get('device_replace_note', ''),
        status=data.get('status', '未处理'),
        fault_cause=data.get('fault_cause', ''),
        solution=data.get('solution', ''),
        completion_time=data.get('completion_time', ''),
        final_status=data.get('final_status', ''),
        photo_url=data.get('photo_url', ''),
        new_classroom=data.get('new_classroom', ''),
        notes=data.get('notes', ''),
        note_images=json.dumps(data.get('note_images', []), ensure_ascii=False),
        created_at=now,
        updated_at=now,
    )
    logger.info(f"创建报修记录: ID={repair.id}, 教室={repair.classroom}")
    return repair.to_dict()


def update_repair(repair_id: int, data: dict) -> dict:
    """更新报修记录"""
    try:
        repair = Repair.get_by_id(repair_id)
    except Repair.DoesNotExist:
        return {'error': '记录不存在'}

    # 管理员可编辑的字段白名单
    ADMIN_EDITABLE_FIELDS = [
        'classroom', 'report_time', 'week_number', 'fault_type', 'reporter_name',
        'reporter_college', 'is_external_teacher', 'report_method', 'handler_name',
        'is_device_replaced', 'device_replace_note', 'status', 'fault_cause',
        'solution', 'completion_time', 'final_status', 'new_classroom', 'notes', 'note_images',
    ]

    update_fields = {}
    for key, value in data.items():
        if key in ADMIN_EDITABLE_FIELDS:
            update_fields[key] = value

    # note_images 需要序列化为 JSON 字符串
    if 'note_images' in update_fields:
        imgs = update_fields['note_images']
        update_fields['note_images'] = json.dumps(imgs, ensure_ascii=False) if isinstance(imgs, list) else imgs
        # 清理被移除的图片文件
        _cleanup_removed_images(repair.note_images or '', update_fields['note_images'])

    update_fields['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    Repair.update(**update_fields).where(Repair.id == repair_id).execute()
    logger.info(f"更新报修记录: ID={repair_id}")
    return Repair.get_by_id(repair_id).to_dict()


# 学生可编辑的字段
STUDENT_EDITABLE_FIELDS = [
    'classroom', 'report_time', 'week_number', 'fault_type', 'reporter_name',
    'reporter_college', 'is_external_teacher', 'report_method', 'handler_name',
    'is_device_replaced', 'device_replace_note', 'status', 'fault_cause',
    'solution', 'notes', 'note_images',
]


def student_update_repair(repair_id: int, student_id: str, data: dict, student_name: str = '') -> dict:
    """学生编辑自己的报修记录（带修改日志）"""
    try:
        repair = Repair.get_by_id(repair_id)
    except Repair.DoesNotExist:
        return {'error': '记录不存在'}

    # 验证是否是该学生的记录（student_id 必须匹配）
    if repair.student_id != student_id:
        return {'error': '无权修改此记录'}

    # 只允许编辑"未处理"状态的记录
    if repair.status != '未处理':
        return {'error': '只能修改状态为"未处理"的记录'}

    # 记录修改日志
    log_entry = {
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'changes': []
    }
    for key in STUDENT_EDITABLE_FIELDS:
        if key in data:
            old_val = str(getattr(repair, key, '') or '')
            new_val = str(data[key] or '')
            if old_val != new_val:
                log_entry['changes'].append({
                    'field': key,
                    'old_value': old_val,
                    'new_value': new_val,
                })

    # 更新字段
    update_fields = {}
    for key in STUDENT_EDITABLE_FIELDS:
        if key in data:
            update_fields[key] = data[key]

    # note_images 需要序列化为 JSON 字符串
    if 'note_images' in update_fields:
        import json as _json
        imgs = update_fields['note_images']
        update_fields['note_images'] = _json.dumps(imgs, ensure_ascii=False) if isinstance(imgs, list) else imgs
        # 清理被移除的图片文件
        _cleanup_removed_images(repair.note_images or '', update_fields['note_images'])

    update_fields['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if update_fields:
        Repair.update(**update_fields).where(Repair.id == repair_id).execute()

    # 保存修改日志
    if log_entry['changes']:
        import json
        ModificationLog.create(
            repair=repair_id,
            change_time=log_entry['time'],
            changes=json.dumps(log_entry['changes'], ensure_ascii=False),
        )
        logger.info(f"学生编辑报修记录: ID={repair_id}, 学生={student_id}, 变更{len(log_entry['changes'])}项")

    return Repair.get_by_id(repair_id).to_dict()


def student_delete_repair(repair_id: int, student_id: str, student_name: str = '') -> dict:
    """学生删除自己的报修记录"""
    try:
        repair = Repair.get_by_id(repair_id)
    except Repair.DoesNotExist:
        return {'error': '记录不存在'}

    # 验证是否是该学生的记录（student_id 匹配 或 handler_name 匹配）
    is_owner = (repair.student_id == student_id) or (student_name and repair.handler_name == student_name)
    if not is_owner:
        return {'error': '无权删除此记录'}

    # 只允许删除"未处理"状态的记录
    if repair.status != '未处理':
        return {'error': '只能删除状态为"未处理"的记录'}

    Repair.delete().where(Repair.id == repair_id).execute()
    logger.info(f"学生删除报修记录: ID={repair_id}, 学生={student_id}")
    return {'success': True}


def delete_repair(repair_id: int) -> bool:
    """删除报修记录"""
    deleted = Repair.delete().where(Repair.id == repair_id).execute()
    if deleted > 0:
        logger.info(f"删除报修记录: ID={repair_id}")
        return True
    return False


def get_repair_by_id(repair_id: int) -> Optional[dict]:
    """根据 ID 获取报修记录"""
    repair = Repair.get_or_none(Repair.id == repair_id)
    return repair.to_dict() if repair else None


def get_student_repairs(student_id: str = '', student_name: str = '',
                        status: str = 'all', keyword: str = '',
                        page: int = 1, page_size: int = 20) -> dict:
    """
    获取学生的报修记录
    支持按 student_id 或 student_name（handler_name）筛选
    """
    query = Repair.select()

    # 按 student_id 或 handler_name 筛选
    if student_id and student_name:
        query = query.where(
            (Repair.student_id == student_id) |
            (Repair.handler_name == student_name)
        )
    elif student_id:
        query = query.where(Repair.student_id == student_id)
    elif student_name:
        query = query.where(Repair.handler_name == student_name)

    # 按状态筛选
    if status and status != 'all':
        query = query.where(Repair.status == status)

    # 按关键词搜索
    if keyword:
        kw = keyword.lower()
        query = query.where(
            (Repair.classroom.contains(kw)) |
            (Repair.fault_type.contains(kw)) |
            (Repair.reporter_name.contains(kw)) |
            (Repair.fault_cause.contains(kw))
        )

    total = query.count()
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))

    records = query.order_by(Repair.report_time.desc()).paginate(page, page_size)
    return {
        'records': [r.to_dict() for r in records],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': total_pages,
    }


# ============================================================
# 智能填充
# ============================================================

def auto_fill(classroom: str, cache, weekday: str = '', section: str = '') -> dict:
    """
    根据教室、星期、节次自动填充报修信息
    :param classroom: 教室名称
    :param cache: 数据缓存对象
    :param weekday: 星期几 (1-7)，可选
    :param section: 节次 (如 "5-6节")，可选
    :return: 自动填充的数据
    """
    import services.admin_config as admin_config

    week_number = admin_config.get_current_week()

    result = {
        'classroom': classroom,
        'report_time': datetime.now().strftime('%Y-%m-%d'),
        'week_number': week_number,
        'course_info': None,
    }

    # 查询课程信息
    try:
        df = cache.get_df()
        if df.empty:
            logger.info("课表数据为空，无法匹配课程")
            return result

        # 清理教室名称（去除空格、统一格式）
        classroom_clean = classroom.strip().replace(' ', '').lower()

        # 如果没有传入星期和节次，使用当前时间
        if not weekday:
            from utils.time_helper import get_auto_time_info
            time_info = get_auto_time_info()
            weekday = str(time_info.get('weekday', ''))
        if not section:
            from utils.time_helper import get_auto_time_info
            time_info = get_auto_time_info()
            section = time_info.get('current_section', '')

        # 策略1：精确匹配（星期 + 节次 + 教室）
        if weekday and section:
            mask = (df['_weekday_str'] == weekday) & \
                   (df['_section_lower'] == section.lower()) & \
                   (df['_classroom_lower'].str.contains(classroom_clean, na=False))

            matched = df[mask]
            if not matched.empty:
                row = matched.iloc[0]
                result['course_info'] = {
                    'course_name': row.get('课程名称', ''),
                    'teacher_name': row.get('姓名', ''),
                    'teacher_id': row.get('教工号', ''),
                    'teacher_college': row.get('开课学院', ''),
                    'class_name': row.get('教学班组成', ''),
                }
                logger.info(f"精确匹配到课程: {result['course_info']['course_name']}")
                return result

        # 策略2：匹配该星期在该教室的所有课程
        if weekday:
            mask = (df['_weekday_str'] == weekday) & \
                   (df['_classroom_lower'].str.contains(classroom_clean, na=False))

            matched = df[mask]
            if not matched.empty:
                row = matched.iloc[0]
                result['course_info'] = {
                    'course_name': row.get('课程名称', ''),
                    'teacher_name': row.get('姓名', ''),
                    'teacher_id': row.get('教工号', ''),
                    'teacher_college': row.get('开课学院', ''),
                    'class_name': row.get('教学班组成', ''),
                }
                logger.info(f"匹配到该日课程: {result['course_info']['course_name']}")
                return result

        # 策略3：匹配任意时间在该教室的课程
        mask = df['_classroom_lower'].str.contains(classroom_clean, na=False)
        matched = df[mask]
        if not matched.empty:
            row = matched.iloc[0]
            result['course_info'] = {
                'course_name': row.get('课程名称', ''),
                'teacher_name': row.get('姓名', ''),
                'teacher_id': row.get('教工号', ''),
                'teacher_college': row.get('开课学院', ''),
                'class_name': row.get('教学班组成', ''),
            }
            logger.info(f"匹配到历史课程: {result['course_info']['course_name']}")

    except Exception as e:
        logger.error(f"查询课程信息失败: {e}")

    return result


# ============================================================
# 就近空教室推荐
# ============================================================

def get_nearby_rooms(classroom: str, weekday: int, section: str, cache) -> dict:
    """
    获取同楼栋的空教室推荐
    :param classroom: 故障教室名称
    :param weekday: 星期几 (1-7)
    :param section: 节次 (如 "5-6节")
    :param cache: 数据缓存对象
    :return: 推荐的空教室列表
    """
    # 提取楼栋和楼层（使用与查询系统一致的 extract_building_name 逻辑）
    from datasource.data_cleaning import extract_building_name
    building = extract_building_name(classroom)
    floor_match = re.search(r'(\d)', classroom)
    fault_floor = int(floor_match.group(1)) if floor_match else 0

    # 查询空教室
    try:
        from datasource.empty_classroom_query import QueryCondition, ClassroomType
        query_system = cache.get_query_system()

        if not query_system:
            return {'fault_room': classroom, 'fault_building': building,
                    'fault_floor': fault_floor, 'recommendations': []}

        # 解析节次并展开（如 "5-6节" → [5, 6]，与空教室查询逻辑一致）
        section_num = _parse_section_number(section)
        expanded_sections = [section_num]
        if section_num % 2 == 1:
            expanded_sections.append(section_num + 1)

        # 模糊匹配楼栋（与空教室查询接口一致的兜底逻辑）
        matched_building = None
        available_buildings = query_system.get_buildings()
        if building in available_buildings:
            matched_building = building
        else:
            for b in available_buildings:
                if building in b or b in building:
                    matched_building = b
                    break

        condition = QueryCondition(
            weekday=weekday,
            sections=sorted(expanded_sections),
            building=matched_building,  # 只查同楼栋
            classroom_type=ClassroomType.ALL,
            exclude_special=True
        )

        results = query_system.query_empty_classrooms(condition)

        # 整理推荐结果
        recommendations = []
        for room in results:
            if room.classroom == classroom:
                continue  # 跳过故障教室本身

            # 提取楼层
            room_floor_match = re.search(r'(\d)', room.classroom.replace(building, ''))
            room_floor = int(room_floor_match.group(1)) if room_floor_match else 0

            # 计算优先级
            if room_floor == fault_floor:
                priority = 1
                tag = '同层'
            else:
                priority = 2
                tag = '同楼'

            recommendations.append({
                'classroom': room.classroom,
                'building': building,
                'floor': room_floor,
                'capacity': 60,  # 默认座位数
                'type': room.classroom_type.value if hasattr(room, 'classroom_type') else '普通',
                'priority': priority,
                'tag': tag,
            })

        # 排序：同层优先，然后按楼层距离排序
        recommendations.sort(key=lambda x: (x['priority'], abs(x['floor'] - fault_floor)))

        return {
            'fault_room': classroom,
            'fault_building': building,
            'fault_floor': fault_floor,
            'recommendations': recommendations,
        }

    except Exception as e:
        logger.error(f"获取空教室推荐失败: {e}")
        return {'fault_room': classroom, 'fault_building': building,
                'fault_floor': fault_floor, 'recommendations': []}


def _parse_section_number(section: str) -> int:
    """解析节次字符串为起始节次数字"""
    section_map = {
        '1-2节': 1, '3-4节': 3, '5-6节': 5,
        '7-8节': 7, '9-10节': 9, '11-12节': 11,
    }
    if section in section_map:
        return section_map[section]
    match = re.match(r'(\d+)', section)
    return int(match.group(1)) if match else 1


# ============================================================
# 报修统计
# ============================================================

def get_dashboard_stats(range_type: str = 'semester') -> dict:
    """
    获取大屏综合统计数据（使用数据库聚合查询优化性能）
    :param range_type: week=本周, month=本月, semester=本学期
    """
    from peewee import fn, Case
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    # 根据时间范围确定起始日期
    if range_type == 'week':
        week_start = (now - timedelta(days=now.weekday())).strftime('%Y-%m-%d')
        range_label = '本周'
        date_filter = Repair.report_time >= week_start
    elif range_type == 'month':
        month_start = now.strftime('%Y-%m-01')
        range_label = '本月'
        date_filter = Repair.report_time >= month_start
    else:
        range_type = 'semester'
        range_label = '本学期'
        date_filter = True  # 无日期过滤

    # 使用数据库聚合查询获取核心指标
    stats = Repair.select(
        fn.COUNT(Repair.id).alias('total'),
        fn.SUM(Case(None, ((Repair.status.in_(['未处理', '处理中']), 1),), 0)).alias('pending'),
        fn.SUM(Case(None, ((Repair.status.in_(['已处理', '已解决']), 1),), 0)).alias('resolved'),
    ).where(date_filter).dicts().get()

    total_count = stats['total'] or 0
    pending_count = stats['pending'] or 0
    resolved_count = stats['resolved'] or 0
    resolved_rate = round(resolved_count / total_count * 100, 1) if total_count > 0 else 0

    # 平均处理天数（使用数据库查询）
    avg_process_days = 0
    try:
        avg_result = Repair.select(
            fn.AVG(
                fn.JULIANDAY(Repair.updated_at) - fn.JULIANDAY(Repair.created_at)
            ).alias('avg_days')
        ).where(
            date_filter & Repair.status.in_(['已处理', '已解决'])
        ).scalar()
        if avg_result:
            avg_process_days = round(float(avg_result), 1)
    except Exception:
        pass

    # 报修趋势（使用数据库聚合）
    import services.admin_config as admin_config
    current_week = admin_config.get_current_week()
    semester_config = admin_config.get_semester_config()

    if range_type == 'week':
        # 按天，从本周一开始到本周日
        monday = now - timedelta(days=now.weekday())  # 本周一
        trend = {}
        for i in range(7):
            d = (monday + timedelta(days=i)).strftime('%Y-%m-%d')
            count = Repair.select(fn.COUNT(Repair.id)).where(
                Repair.report_time.startswith(d)
            ).scalar() or 0
            trend[d] = count
    elif range_type == 'month':
        # 按天，30天
        trend = {}
        for i in range(30):
            d = (now - timedelta(days=29 - i)).strftime('%Y-%m-%d')
            count = Repair.select(fn.COUNT(Repair.id)).where(
                Repair.report_time.startswith(d)
            ).scalar() or 0
            trend[d] = count
    else:
        # 按教学周聚合
        start_date_str = semester_config.get('start_date', '')
        start_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        trend = {}
        if start_date:
            # 使用数据库查询获取每周数据
            repairs = Repair.select(Repair.report_time).where(date_filter)
            for r in repairs:
                rt = r.report_time[:10] if r.report_time else ''
                if rt:
                    try:
                        dt = datetime.strptime(rt, '%Y-%m-%d').date()
                        days_diff = (dt - start_date).days
                        if days_diff >= 0:
                            week_num = days_diff // 7 + 1
                            if week_num <= current_week:
                                week_label = f"第{week_num}周"
                                trend[week_label] = trend.get(week_label, 0) + 1
                    except ValueError:
                        pass

            # 确保所有周都有数据（填充0）
            for w in range(1, current_week + 1):
                key = f"第{w}周"
                if key not in trend:
                    trend[key] = 0

        # 按周号数字排序
        def _week_sort_key(item):
            try:
                return int(item[0].replace('第', '').replace('周', ''))
            except ValueError:
                return 0
        trend = dict(sorted(trend.items(), key=_week_sort_key))

    # 故障类型分布（使用数据库聚合）
    fault_types = {}
    for ft in Repair.select(Repair.fault_type, fn.COUNT(Repair.id).alias('count')).where(
        date_filter
    ).group_by(Repair.fault_type):
        key = ft.fault_type or '其他'
        fault_types[key] = ft.count

    # 楼栋分布（使用数据库聚合）
    building_stats = {}
    for r in Repair.select(Repair.classroom, fn.COUNT(Repair.id).alias('count')).where(
        date_filter
    ).group_by(Repair.classroom):
        room = r.classroom or ''
        building = re.sub(r'\d+$', '', room).strip()
        if building:
            building_stats[building] = building_stats.get(building, 0) + r.count
    building_stats = dict(sorted(building_stats.items(), key=lambda x: x[1], reverse=True)[:10])

    # 学院分布（使用数据库聚合）
    college_stats = {}
    for r in Repair.select(Repair.reporter_college, fn.COUNT(Repair.id).alias('count')).where(
        date_filter & (Repair.reporter_college != '')
    ).group_by(Repair.reporter_college):
        college = r.reporter_college.strip() if r.reporter_college else ''
        if college:
            college_stats[college] = r.count
    college_stats = dict(sorted(college_stats.items(), key=lambda x: x[1], reverse=True))

    # 处理人排行（使用数据库聚合）
    handler_stats = {}
    for r in Repair.select(Repair.handler_name, fn.COUNT(Repair.id).alias('count')).where(
        date_filter & (Repair.handler_name != '')
    ).group_by(Repair.handler_name):
        handler = r.handler_name.strip() if r.handler_name else ''
        if handler:
            handler_stats[handler] = r.count
    handler_stats = dict(sorted(handler_stats.items(), key=lambda x: x[1], reverse=True))

    # 星期分布（使用数据库查询）
    weekday_stats = {}
    repairs = Repair.select(Repair.report_time).where(date_filter)
    for r in repairs:
        rt = r.report_time[:10] if r.report_time else ''
        if rt:
            try:
                dt = datetime.strptime(rt, '%Y-%m-%d')
                wd = str(dt.weekday() + 1)
                weekday_stats[wd] = weekday_stats.get(wd, 0) + 1
            except ValueError:
                pass

    # 最新报修动态（取最新12条，不受时间范围限制）
    latest = [r.to_dict_light() for r in Repair.select().order_by(Repair.report_time.desc()).limit(12)]

    return {
        'range_type': range_type,
        'range_label': range_label,
        'total_count': total_count,
        'pending_count': pending_count,
        'resolved_count': resolved_count,
        'resolved_rate': resolved_rate,
        'avg_process_days': avg_process_days,
        'trend': trend,
        'fault_types': fault_types,
        'building_stats': building_stats,
        'college_stats': college_stats,
        'handler_stats': handler_stats,
        'weekday_stats': weekday_stats,
        'latest_repairs': latest,
        'current_week': current_week,
        'total_weeks': 20,
        'semester_start': semester_config.get('start_date', ''),
    }


def get_building_drill(building: str, range_type: str = 'semester') -> dict:
    """楼栋下钻统计"""
    ranged = _to_light_dicts(
        Repair.select().where(
            Repair.classroom.contains(building) & _sql_date_filter(range_type)
        )
    )

    total = len(ranged)
    pending = len([r for r in ranged if r.get('status') in ('未处理', '处理中')])
    resolved = len([r for r in ranged if r.get('status') in ('已处理', '已解决')])
    rate = round(resolved / total * 100, 1) if total > 0 else 0

    # 最常故障教室
    room_counts = count_dict(ranged, 'classroom', 1)
    most_room = list(room_counts.keys())[0] if room_counts else '--'

    # 故障类型分布
    fault_types = count_dict(ranged, 'fault_type')

    # 楼层分布
    floor_stats = {}
    for r in ranged:
        room = r.get('classroom', '')
        num = room.replace(building, '').strip()
        m = re.match(r'(\d)', num)
        floor = f"{m.group(1)}楼" if m else "其他"
        floor_stats[floor] = floor_stats.get(floor, 0) + 1
    floor_stats = dict(sorted(floor_stats.items()))

    # 教室排行
    room_stats = count_dict(ranged, 'classroom', 10)

    # 趋势
    trend = trend_data(ranged, range_type)

    # 最近记录
    latest = sorted(ranged, key=lambda x: x.get('report_time', ''), reverse=True)[:10]

    return {
        'building': building,
        'range_type': range_type,
        'total': total, 'pending': pending, 'resolved': resolved, 'rate': rate,
        'most_room': most_room,
        'fault_types': fault_types,
        'floor_stats': floor_stats,
        'room_stats': room_stats,
        'trend': trend,
        'latest': latest,
    }


def get_college_drill(college: str, range_type: str = 'semester') -> dict:
    """学院下钻统计"""
    ranged = _to_light_dicts(
        Repair.select().where(
            (Repair.reporter_college == college) & _sql_date_filter(range_type)
        )
    )

    total = len(ranged)
    pending = len([r for r in ranged if r.get('status') in ('未处理', '处理中')])
    resolved = len([r for r in ranged if r.get('status') in ('已处理', '已解决')])
    rate = round(resolved / total * 100, 1) if total > 0 else 0

    # 报修教师数
    teacher_count = len(set(r.get('reporter_name', '') for r in ranged if r.get('reporter_name')))

    # 教师排行
    teacher_stats = count_dict(ranged, 'reporter_name', 10)

    # 故障类型
    fault_types = count_dict(ranged, 'fault_type')

    # 教室分布
    room_stats = count_dict(ranged, 'classroom', 10)

    # 报修方式
    method_stats = count_dict(ranged, 'report_method')

    # 趋势
    trend = trend_data(ranged, range_type)

    latest = sorted(ranged, key=lambda x: x.get('report_time', ''), reverse=True)[:10]

    return {
        'college': college, 'range_type': range_type,
        'total': total, 'pending': pending, 'resolved': resolved, 'rate': rate,
        'teacher_count': teacher_count,
        'teacher_stats': teacher_stats,
        'fault_types': fault_types,
        'room_stats': room_stats,
        'method_stats': method_stats,
        'trend': trend,
        'latest': latest,
    }


def get_handler_drill(handler: str, range_type: str = 'semester') -> dict:
    """处理人下钻统计"""
    ranged = _to_light_dicts(
        Repair.select().where(
            (Repair.handler_name == handler) & _sql_date_filter(range_type)
        )
    )

    total = len(ranged)
    pending = len([r for r in ranged if r.get('status') in ('未处理', '处理中')])
    resolved = len([r for r in ranged if r.get('status') in ('已处理', '已解决')])
    rate = round(resolved / total * 100, 1) if total > 0 else 0
    avg_days = avg_process_days(ranged)

    # 系统平均处理天数（用于对比，单独查询）
    sys_avg = avg_process_days(_to_light_dicts(
        Repair.select().where(Repair.status.in_(['已处理', '已解决']))
    ))

    # 故障类型
    fault_types = count_dict(ranged, 'fault_type')

    # 楼栋分布
    building_stats = {}
    for r in ranged:
        building = re.sub(r'\d+$', '', r.get('classroom', '')).strip()
        if building:
            building_stats[building] = building_stats.get(building, 0) + 1
    building_stats = dict(sorted(building_stats.items(), key=lambda x: x[1], reverse=True)[:10])

    # 趋势
    trend = trend_data(ranged, range_type)

    # 待处理列表
    pending_list = [r for r in ranged if r.get('status') in ('未处理', '处理中')]
    pending_list.sort(key=lambda x: x.get('report_time', ''), reverse=True)

    return {
        'handler': handler, 'range_type': range_type,
        'total': total, 'pending': pending, 'resolved': resolved, 'rate': rate,
        'avg_days': avg_days, 'sys_avg_days': sys_avg,
        'fault_types': fault_types,
        'building_stats': building_stats,
        'trend': trend,
        'pending_list': pending_list,
    }


def get_fault_type_drill(fault_type: str, range_type: str = 'semester') -> dict:
    """故障类型下钻统计"""
    date_cond = _sql_date_filter(range_type)
    ranged = _to_light_dicts(
        Repair.select().where(
            (Repair.fault_type == fault_type) & date_cond
        )
    )

    total = len(ranged)
    all_total = Repair.select().where(date_cond).count()
    ratio = round(total / all_total * 100, 1) if all_total > 0 else 0

    # 涉及楼栋数和教室数
    buildings = set()
    classrooms = set()
    for r in ranged:
        room = r.get('classroom', '')
        if room:
            classrooms.add(room)
            building = re.sub(r'\d+$', '', room).strip()
            if building:
                buildings.add(building)

    # 楼栋分布
    building_stats = {}
    for r in ranged:
        building = re.sub(r'\d+$', '', r.get('classroom', '')).strip()
        if building:
            building_stats[building] = building_stats.get(building, 0) + 1
    building_stats = dict(sorted(building_stats.items(), key=lambda x: x[1], reverse=True)[:10])

    # 学院分布
    college_stats = count_dict(ranged, 'reporter_college', 10)

    # 处理人分布
    handler_stats = count_dict(ranged, 'handler_name')

    # 趋势
    trend = trend_data(ranged, range_type)

    latest = sorted(ranged, key=lambda x: x.get('report_time', ''), reverse=True)[:10]

    return {
        'fault_type': fault_type, 'range_type': range_type,
        'total': total, 'ratio': ratio,
        'building_count': len(buildings), 'classroom_count': len(classrooms),
        'building_stats': building_stats,
        'college_stats': college_stats,
        'handler_stats': handler_stats,
        'trend': trend,
        'latest': latest,
    }


def get_repair_detail(repair_id: int) -> dict:
    """单条报修详情（含同教室历史）"""
    try:
        target_obj = Repair.get_by_id(repair_id)
    except Repair.DoesNotExist:
        return {'error': '记录不存在'}

    target = target_obj.to_dict()  # 含修改日志
    classroom = target.get('classroom', '')

    # 同教室历史（轻量字典，无修改日志）
    same_room = _to_light_dicts(
        Repair.select().where(
            (Repair.classroom == classroom) & (Repair.id != repair_id)
        ).order_by(Repair.report_time.desc()).limit(10)
    )

    # 该教室故障类型分布
    room_all = _to_light_dicts(
        Repair.select().where(Repair.classroom == classroom)
    )
    room_fault_types = count_dict(room_all, 'fault_type')

    return {
        'repair': target,
        'same_room_history': same_room,
        'classroom': classroom,
        'room_fault_types': room_fault_types,
    }


def get_pending_list(range_type: str = 'semester') -> dict:
    """获取所有待处理工单"""
    pending = _to_light_dicts(
        Repair.select().where(
            Repair.status.in_(['未处理', '处理中']) & _sql_date_filter(range_type)
        ).order_by(Repair.report_time.desc())
    )

    # 按状态分组
    unhandled = [r for r in pending if r.get('status') == '未处理']
    processing = [r for r in pending if r.get('status') == '处理中']

    # 按楼栋分组
    building_stats = {}
    for r in pending:
        building = re.sub(r'\d+$', '', r.get('classroom', '')).strip()
        if building:
            building_stats[building] = building_stats.get(building, 0) + 1
    building_stats = dict(sorted(building_stats.items(), key=lambda x: x[1], reverse=True))

    # 按故障类型分组
    fault_stats = count_dict(pending, 'fault_type')

    return {
        'total': len(pending),
        'unhandled_count': len(unhandled),
        'processing_count': len(processing),
        'pending_list': pending,
        'building_stats': building_stats,
        'fault_stats': fault_stats,
    }


def get_classroom_drill(classroom: str, range_type: str = 'semester') -> dict:
    """教室下钻统计"""
    date_cond = _sql_date_filter(range_type)
    ranged = _to_light_dicts(
        Repair.select().where(
            (Repair.classroom == classroom) & date_cond
        )
    )

    total = len(ranged)
    if total == 0:
        # 尝试模糊匹配
        ranged = _to_light_dicts(
            Repair.select().where(
                Repair.classroom.contains(classroom) & date_cond
            )
        )
        total = len(ranged)

    pending = len([r for r in ranged if r.get('status') in ('未处理', '处理中')])
    resolved = len([r for r in ranged if r.get('status') in ('已处理', '已解决')])
    rate = round(resolved / total * 100, 1) if total > 0 else 0
    avg_days = avg_process_days(ranged)

    # 最常出现的问题（故障类型）
    fault_stats = count_dict(ranged, 'fault_type')

    # 最常出现的故障原因
    cause_stats = {}
    for r in ranged:
        cause = r.get('fault_cause', '').strip()
        if cause and len(cause) > 1:
            # 截取前20个字符作为分组key
            short = cause[:20]
            cause_stats[short] = cause_stats.get(short, 0) + 1
    cause_stats = dict(sorted(cause_stats.items(), key=lambda x: x[1], reverse=True)[:8])

    # 处理人分布
    handler_stats = count_dict(ranged, 'handler_name')

    # 报修人分布
    reporter_stats = count_dict(ranged, 'reporter_name', 8)

    # 报修趋势
    trend = trend_data(ranged, range_type)

    return {
        'classroom': classroom,
        'range_type': range_type,
        'total': total, 'pending': pending, 'resolved': resolved, 'rate': rate,
        'avg_days': avg_days,
        'fault_stats': fault_stats,
        'cause_stats': cause_stats,
        'handler_stats': handler_stats,
        'reporter_stats': reporter_stats,
        'trend': trend,
        'latest': sorted(ranged, key=lambda x: x.get('report_time', ''), reverse=True)[:10],
    }


def get_repair_stats() -> dict:
    """获取报修统计数据"""
    from peewee import fn, Case
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    # 核心指标（单条 SQL 聚合）
    stats = Repair.select(
        fn.COUNT(Repair.id).alias('total'),
        fn.SUM(Case(None, ((Repair.report_time.startswith(today), 1),), 0)).alias('today'),
        fn.SUM(Case(None, ((Repair.status.in_(['未处理', '处理中']), 1),), 0)).alias('pending'),
        fn.SUM(Case(None, ((Repair.status.in_(['已处理', '已解决']), 1),), 0)).alias('resolved'),
    ).dicts().get()

    total_count = stats['total'] or 0
    today_count = stats['today'] or 0
    pending_count = stats['pending'] or 0
    resolved_count = stats['resolved'] or 0

    # 故障类型分布
    fault_types = {}
    for r in Repair.select(Repair.fault_type, fn.COUNT(Repair.id).alias('count')).group_by(Repair.fault_type):
        fault_types[r.fault_type or '其他'] = r.count

    # 近7天报修趋势
    trend = {}
    for i in range(7):
        date = (now - timedelta(days=6-i)).strftime('%Y-%m-%d')
        trend[date] = Repair.select(fn.COUNT(Repair.id)).where(
            Repair.report_time.startswith(date)
        ).scalar() or 0

    # 楼栋分布
    building_stats = {}
    for r in Repair.select(Repair.classroom, fn.COUNT(Repair.id).alias('count')).group_by(Repair.classroom):
        room = r.classroom or ''
        building = re.sub(r'\d+$', '', room).strip()
        if building:
            building_stats[building] = building_stats.get(building, 0) + r.count

    return {
        'today_count': today_count,
        'pending_count': pending_count,
        'resolved_count': resolved_count,
        'total_count': total_count,
        'fault_types': fault_types,
        'trend': trend,
        'building_stats': building_stats,
    }


# ============================================================
# Excel 导出
# ============================================================

def export_to_excel(range_type: str = 'all',
                    semester: str = '',
                    start_date: str = '',
                    end_date: str = '',
                    status: str = 'all'):
    """
    导出报修记录为 Excel 文件（返回内存缓冲区，不保存到磁盘）
    :return: (BytesIO缓冲区, 文件名)
    """
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')

    # 计算时间范围
    if range_type == 'today':
        start_date = today
        end_date = today
    elif range_type == 'week':
        monday = now - timedelta(days=now.weekday())
        start_date = monday.strftime('%Y-%m-%d')
        end_date = today
    elif range_type == 'month':
        start_date = now.strftime('%Y-%m-01')
        end_date = today

    # 获取数据（page_size=99999 确保导出全部记录）
    result = get_repair_list(
        status=status,
        semester=semester,
        start_date=start_date,
        end_date=end_date,
        page=1,
        page_size=99999
    )
    records = result.get('records', []) if isinstance(result, dict) else result

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '工作表1'

    # 标题行
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    if semester:
        title_cell.value = f'綦江校区{semester}设备信息及报修日志记录表'
    else:
        title_cell.value = f'綦江校区{now.year}年上半年设备信息及报修日志记录表'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = [
        '序号', '故障教室/书院', '报修时间', '周次', '报修类型',
        '报修人\n（不允许填未知）', '报修人学院', '是否为外聘教师',
        '报修方式', '处理人', '是否更换设备', '设备更换备注', '处理情况',
        '故障发生原因', '处理方式、或未处理原因', '若未处理，处理完毕时间', '最终状态'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 数据行
    for i, record in enumerate(records, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)  # 序号
        ws.cell(row=row, column=2, value=record.get('classroom', ''))
        ws.cell(row=row, column=3, value=record.get('report_time', ''))
        ws.cell(row=row, column=4, value=record.get('week_number', ''))
        ws.cell(row=row, column=5, value=record.get('fault_type', ''))
        ws.cell(row=row, column=6, value=record.get('reporter_name', ''))
        ws.cell(row=row, column=7, value=record.get('reporter_college', ''))
        ws.cell(row=row, column=8, value='是' if record.get('is_external_teacher') else '否')
        ws.cell(row=row, column=9, value=record.get('report_method', ''))
        ws.cell(row=row, column=10, value=record.get('handler_name', ''))
        ws.cell(row=row, column=11, value='是' if record.get('is_device_replaced') else '否')
        ws.cell(row=row, column=12, value=record.get('device_replace_note', ''))
        ws.cell(row=row, column=13, value=record.get('status', ''))
        ws.cell(row=row, column=14, value=record.get('fault_cause', ''))
        ws.cell(row=row, column=15, value=record.get('solution', ''))
        ws.cell(row=row, column=16, value=record.get('completion_time', ''))
        ws.cell(row=row, column=17, value=record.get('final_status', ''))

    # 设置列宽
    col_widths = [8, 15, 12, 8, 12, 15, 18, 12, 15, 12, 12, 20, 12, 25, 30, 20, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 保存到内存缓冲区（不写磁盘）
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'报修记录_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'

    logger.info(f"导出 Excel: {filename}, 共 {len(records)} 条记录")
    return buffer, filename


# ============================================================
# Excel 导入
# ============================================================

# Excel 列名到字段名的映射
COLUMN_MAPPING = {
    '序号': '_index',
    '故障教室/书院': 'classroom',
    '故障教室': 'classroom',
    '教室': 'classroom',
    '报修时间': 'report_time',
    '周次': 'week_number',
    '报修类型': 'fault_type',
    '报修人': 'reporter_name',
    '报修人（不允许填未知）': 'reporter_name',
    '报修人学院': 'reporter_college',
    '学院': 'reporter_college',
    '是否为外聘教师': 'is_external_teacher',
    '外聘教师': 'is_external_teacher',
    '报修方式': 'report_method',
    '处理人': 'handler_name',
    '是否更换设备': 'is_device_replaced',
    '处理情况': 'status',
    '故障发生原因': 'fault_cause',
    '故障原因': 'fault_cause',
    '处理方式、或未处理原因': 'solution',
    '处理方式': 'solution',
    '若未处理，处理完毕时间': 'completion_time',
    '处理完毕时间': 'completion_time',
}

# 必填字段
REQUIRED_FIELDS = ['classroom', 'fault_type', 'reporter_name', 'handler_name']

# 字段类型定义
FIELD_TYPES = {
    'classroom': str,
    'report_time': str,
    'week_number': int,
    'fault_type': str,
    'reporter_name': str,
    'reporter_college': str,
    'is_external_teacher': bool,
    'report_method': str,
    'handler_name': str,
    'is_device_replaced': bool,
    'status': str,
    'fault_cause': str,
    'solution': str,
    'completion_time': str,
    'final_status': str,
}

# 有效的报修类型
VALID_FAULT_TYPES = ['中控', '电脑', '投影仪', '音响', '麦克风', '展台', '幕布', '网络', '软件', '其他']

# 有效的处理状态
VALID_STATUS = ['未处理', '处理中', '已处理', '已解决']


def import_from_excel(filepath: str) -> dict:
    """
    从 Excel 文件导入报修记录
    智能识别列结构，自动适配不同格式的 Excel 文件

    :param filepath: Excel 文件路径
    :return: 导入结果 {success_count, error_count, errors, duplicates}
    """
    result = {
        'success_count': 0,
        'error_count': 0,
        'duplicate_count': 0,
        'errors': [],
        'duplicates': [],
        'imported_records': [],
        'detected_columns': {},
    }

    try:
        # 读取 Excel 文件（跳过第1行标题，第2行是表头）
        df = _read_excel_robust(filepath)

        if df is None or df.empty:
            result['errors'].append({'row': 0, 'message': 'Excel 文件为空或无法读取'})
            return result

        # 智能识别列映射
        column_map = _detect_column_mapping(df)
        result['detected_columns'] = column_map

        if not column_map:
            result['errors'].append({
                'row': 0,
                'message': '无法识别表格列结构，请确保表格包含报修相关字段（如：故障教室、报修时间、报修类型、报修人、处理人等）'
            })
            return result

        # 应用列映射
        df = df.rename(columns=column_map)

        # 加载现有数据的去重键（从数据库查询）
        existing_keys = set()
        for r in Repair.select(Repair.classroom, Repair.report_time, Repair.fault_type):
            key = f"{r.classroom}_{r.report_time}_{r.fault_type}"
            existing_keys.add(key)

        # 逐行处理
        imported_count = 0
        for idx, row in df.iterrows():
            row_num = idx + 3  # Excel 行号（第1行标题，第2行表头，数据从第3行）

            try:
                record = _process_excel_row(row, row_num)
                if record is None:
                    continue

                # 检查重复
                key = f"{record.get('classroom', '')}_{record.get('report_time', '')}_{record.get('fault_type', '')}"
                if key in existing_keys:
                    result['duplicate_count'] += 1
                    result['duplicates'].append({
                        'row': row_num,
                        'classroom': record.get('classroom', ''),
                        'report_time': record.get('report_time', ''),
                    })
                    continue

                # 写入数据库
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                Repair.create(
                    semester=_get_current_semester(),
                    classroom=record.get('classroom', ''),
                    report_time=record.get('report_time', ''),
                    week_number=int(record.get('week_number', 0)),
                    fault_type=record.get('fault_type', ''),
                    reporter_name=record.get('reporter_name', ''),
                    reporter_college=record.get('reporter_college', ''),
                    is_external_teacher=bool(record.get('is_external_teacher', False)),
                    report_method=record.get('report_method', ''),
                    handler_name=record.get('handler_name', ''),
                    is_device_replaced=bool(record.get('is_device_replaced', False)),
                    status=record.get('status', '未处理'),
                    fault_cause=record.get('fault_cause', ''),
                    solution=record.get('solution', ''),
                    completion_time=record.get('completion_time', ''),
                    final_status=record.get('final_status', ''),
                    created_at=now,
                    updated_at=now,
                )

                imported_count += 1
                existing_keys.add(key)
                result['success_count'] += 1

            except Exception as e:
                result['error_count'] += 1
                result['errors'].append({
                    'row': row_num,
                    'message': str(e)
                })

        # 记录导入结果
        if imported_count > 0:
            logger.info(f"Excel 导入完成: 成功 {result['success_count']} 条, "
                       f"重复 {result['duplicate_count']} 条, "
                       f"错误 {result['error_count']} 条")

    except Exception as e:
        logger.error(f"Excel 导入失败: {e}")
        result['errors'].append({'row': 0, 'message': f'文件读取失败: {str(e)}'})

    return result


def _read_excel_robust(filepath: str) -> pd.DataFrame:
    """
    健壮的 Excel 读取函数
    处理 openpyxl 无法读取的带复杂样式的 Excel 文件
    """
    import zipfile
    import xml.etree.ElementTree as ET

    # 方法1: 尝试 openpyxl（标准方式）
    try:
        df = pd.read_excel(filepath, header=1, engine='openpyxl')
        df.columns = df.columns.str.strip()
        if not df.empty:
            return df
    except Exception as e:
        logger.warning(f"openpyxl 读取失败，尝试 XML 解析: {e}")

    # 方法2: 直接解析 xlsx 的 XML（绕过样式问题）
    try:
        return _read_xlsx_via_xml(filepath)
    except Exception as e:
        logger.error(f"XML 解析也失败: {e}")
        raise RuntimeError(f"无法读取 Excel 文件: {e}")


def _read_xlsx_via_xml(filepath: str) -> pd.DataFrame:
    """
    直接解析 xlsx 文件的 XML 内容
    绕过 openpyxl 的样式解析问题
    """
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(filepath, 'r') as z:
        # 读取共享字符串表
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in root.findall('.//ns:si', ns):
                    t = si.find('.//ns:t', ns)
                    if t is not None and t.text:
                        shared_strings.append(t.text)
                    else:
                        shared_strings.append('')

        # 读取工作表
        sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet')]
        if not sheet_files:
            raise RuntimeError("找不到工作表文件")

        with z.open(sheet_files[0]) as f:
            tree = ET.parse(f)
            root = tree.getroot()
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            rows_data = []
            for row in root.findall('.//ns:row', ns):
                cells = {}
                for cell in row.findall('ns:c', ns):
                    ref = cell.get('r', '')
                    # 提取列字母
                    col_letter = ''.join(c for c in ref if c.isalpha())
                    cell_type = cell.get('t', '')
                    v = cell.find('ns:v', ns)
                    val = v.text if v is not None else ''

                    # 解析共享字符串
                    if cell_type == 's' and val:
                        try:
                            idx = int(val)
                            val = shared_strings[idx] if idx < len(shared_strings) else val
                        except (ValueError, IndexError):
                            pass

                    cells[col_letter] = val

                if cells:
                    rows_data.append(cells)

        if len(rows_data) < 3:
            raise RuntimeError("Excel 文件数据不足")

        # 第1行是标题（跳过），第2行是表头，第3行开始是数据
        header_row = rows_data[1]  # 第2行是表头
        data_rows = rows_data[2:]  # 第3行开始是数据

        # 构建列名列表（按列字母排序，排除空列名，处理重复列名）
        col_letters = []
        seen_headers = set()
        for col in sorted(header_row.keys(), key=lambda x: (len(x), x)):
            header = header_row.get(col, '').strip()
            if not header:
                continue
            # 处理重复列名：第一次出现用原名，后续加后缀
            if header in seen_headers:
                # 跳过重复的列（保留第一个）
                continue
            seen_headers.add(header)
            col_letters.append(col)

        # 创建 DataFrame
        records = []
        for row_data in data_rows:
            record = {}
            for col in col_letters:
                header = header_row.get(col, '').strip()
                value = row_data.get(col, '')
                record[header] = value
            records.append(record)

        df = pd.DataFrame(records)
        df.columns = df.columns.str.strip()

        logger.info(f"XML 解析成功: {len(df)} 行, {len(df.columns)} 列")
        return df


def _detect_column_mapping(df: pd.DataFrame) -> dict:
    """
    智能检测 DataFrame 的列映射关系
    通过分析列名和数据模式来判断每列的实际含义
    """
    column_map = {}
    columns = list(df.columns)

    for col in columns:
        col_str = str(col).strip()
        col_lower = col_str.lower()

        # 1. 精确匹配列名
        if col_str in COLUMN_MAPPING:
            column_map[col] = COLUMN_MAPPING[col_str]
            continue

        # 2. 模糊匹配列名
        matched = False
        for key, value in COLUMN_MAPPING.items():
            if key in col_str or col_str in key:
                column_map[col] = value
                matched = True
                break
        if matched:
            continue

        # 3. 基于关键词匹配
        if '教室' in col_str or '书院' in col_str or '地点' in col_str:
            column_map[col] = 'classroom'
        elif '报修时间' in col_str or '时间' in col_str:
            column_map[col] = 'report_time'
        elif '周次' in col_str or '周' == col_str:
            column_map[col] = 'week_number'
        elif '报修类型' in col_str or '类型' in col_str:
            column_map[col] = 'fault_type'
        elif '报修人' in col_str and '学院' not in col_str:
            column_map[col] = 'reporter_name'
        elif '学院' in col_str or '报修人学院' in col_str:
            column_map[col] = 'reporter_college'
        elif '外聘' in col_str:
            column_map[col] = 'is_external_teacher'
        elif '报修方式' in col_str:
            column_map[col] = 'report_method'
        elif '处理人' in col_str:
            column_map[col] = 'handler_name'
        elif '更换设备' in col_str:
            column_map[col] = 'is_device_replaced'
        elif '处理情况' in col_str or '处理状态' in col_str:
            column_map[col] = 'status'
        elif '故障原因' in col_str or '故障发生原因' in col_str:
            column_map[col] = 'fault_cause'
        elif '处理方式' in col_str or '处理方法' in col_str:
            column_map[col] = 'solution'
        elif '完毕时间' in col_str:
            column_map[col] = 'completion_time'

    # 如果列名匹配不到足够字段，尝试基于数据模式推断
    if len(column_map) < 5:
        column_map = _infer_columns_from_data(df)

    return column_map


def _infer_columns_from_data(df: pd.DataFrame) -> dict:
    """
    基于数据模式推断列的含义
    当列名无法匹配时使用
    """
    column_map = {}
    columns = list(df.columns)
    sample_rows = min(10, len(df))

    for i, col in enumerate(columns):
        values = [str(v).strip() for v in df[col].head(sample_rows) if pd.notna(v) and str(v).strip()]

        if not values:
            continue

        # 检查是否是序号列（1,2,3,4,5...）
        if all(v.isdigit() and int(v) < 1000 for v in values):
            if i == 0:  # 第一列通常是序号
                continue

        # 检查是否是日期列（Excel 日期序列号 40000-50000 范围）
        if all(v.replace('.', '').isdigit() and 40000 < float(v) < 60000 for v in values):
            column_map[col] = 'report_time'
            continue

        # 检查是否是周次列（1-30 的小数字）
        if all(v.isdigit() and 1 <= int(v) <= 30 for v in values):
            if 'week_number' not in column_map.values():
                column_map[col] = 'week_number'
                continue

        # 检查是否是教室列（包含楼栋名称）
        building_keywords = ['楼', '教室', '实验室', '机房', '厅']
        if any(any(kw in v for kw in building_keywords) for v in values):
            column_map[col] = 'classroom'
            continue

        # 检查是否是报修类型列
        fault_types = ['中控', '电脑', '投影仪', '音响', '麦克风', '展台', '幕布', '网络', '软件', '话筒', '线路', '云桌面', '电源']
        if any(v in fault_types for v in values):
            column_map[col] = 'fault_type'
            continue

        # 检查是否是状态列
        status_values = ['已处理', '未处理', '处理中', '已解决']
        if any(v in status_values for v in values):
            column_map[col] = 'status'
            continue

        # 检查是否是是否列
        yn_values = ['是', '否']
        if all(v in yn_values for v in values):
            if 'is_external_teacher' not in column_map.values():
                column_map[col] = 'is_external_teacher'
                continue

        # 检查是否是报修方式列
        method_values = ['多媒体报修群', '电话报修', '现场报修', '电话', '微信']
        if any(v in method_values for v in values):
            column_map[col] = 'report_method'
            continue

    return column_map


def _process_excel_row(row, row_num: int) -> dict:
    """处理 Excel 单行数据，支持智能识别和数据清洗"""
    record = {}

    # 提取序号字段（用于判断是否为空行）
    index_val = row.get('_index', '') or row.get('序号', '')
    if pd.isna(index_val):
        index_val = ''
    index_val = str(index_val).strip()

    # 跳过空行（序号列为空）
    if not index_val:
        return None

    # 提取其他字段值
    for field in FIELD_TYPES:
        value = row.get(field, '')
        if pd.isna(value):
            value = ''
        record[field] = value

    # 验证必填字段（放宽限制，只检查教室）
    classroom = str(record.get('classroom', '')).strip()
    if not classroom:
        # 尝试从其他字段找教室
        for field_name, value in record.items():
            val_str = str(value).strip()
            if any(kw in val_str for kw in ['楼', '教室', '实验室', '机房']):
                classroom = val_str
                record['classroom'] = classroom
                break

    if not classroom:
        raise ValueError(f"第{row_num}行: 无法识别教室名称")

    # 数据类型转换和验证
    record['classroom'] = classroom

    # 报修时间处理（支持 Excel 日期序列号）
    report_time = record.get('report_time', '')
    record['report_time'] = _parse_excel_date(report_time)

    # 周次
    week_val = record.get('week_number', '')
    record['week_number'] = _parse_week_number(week_val)

    # 报修类型验证和标准化
    fault_type = str(record.get('fault_type', '')).strip()
    record['fault_type'] = _normalize_fault_type(fault_type)

    # 处理状态验证
    status = str(record.get('status', '')).strip()
    record['status'] = _normalize_status(status)

    # 外聘教师
    ext_val = str(record.get('is_external_teacher', '')).strip()
    record['is_external_teacher'] = ext_val in ('是', 'True', 'true', '1', 'yes')

    # 更换设备
    replace_val = str(record.get('is_device_replaced', '')).strip()
    record['is_device_replaced'] = replace_val in ('是', 'True', 'true', '1', 'yes')

    # 字符串字段清理
    str_fields = ['reporter_name', 'reporter_college', 'report_method',
                  'handler_name', 'fault_cause', 'solution', 'completion_time', 'final_status']
    for field in str_fields:
        val = record.get(field, '')
        record[field] = str(val).strip() if val and not pd.isna(val) else ''

    # 如果缺少报修人，尝试推断
    if not record.get('reporter_name'):
        record['reporter_name'] = '未知'

    # 如果缺少处理人，默认为空
    if not record.get('handler_name'):
        record['handler_name'] = ''

    return record


def _parse_excel_date(value) -> str:
    """解析 Excel 日期（支持序列号和字符串格式）"""
    if not value or (isinstance(value, str) and not value.strip()):
        return datetime.now().strftime('%Y-%m-%d')

    value_str = str(value).strip()

    # 如果是 datetime 对象
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')

    # 如果是 Excel 日期序列号（40000-60000 范围）
    try:
        num = float(value_str)
        if 40000 < num < 60000:
            # Excel 日期序列号转换
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)
            result = base_date + timedelta(days=num)
            return result.strftime('%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        pass

    # 如果是标准日期格式
    date_formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d',
        '%m/%d/%Y',
    ]
    for fmt in date_formats:
        try:
            return datetime.strptime(value_str, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue

    return value_str


def _parse_week_number(value) -> int:
    """解析周次"""
    if not value or (isinstance(value, str) and not value.strip()):
        return 0
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def _normalize_fault_type(value: str) -> str:
    """标准化报修类型"""
    if not value:
        return '其他'

    value = value.strip()

    # 直接匹配
    if value in VALID_FAULT_TYPES:
        return value

    # 模糊匹配
    type_mapping = {
        '中控': '中控',
        '电脑': '电脑',
        '计算机': '电脑',
        '投影': '投影仪',
        '投影仪': '投影仪',
        '音响': '音响',
        '喇叭': '音响',
        '麦克': '麦克风',
        '话筒': '麦克风',
        '麦克风': '麦克风',
        '展台': '展台',
        '幕布': '幕布',
        '网络': '网络',
        '网线': '网络',
        '软件': '软件',
        '线路': '其他',
        '电源': '其他',
        '云桌面': '电脑',
    }

    for keyword, fault_type in type_mapping.items():
        if keyword in value:
            return fault_type

    return '其他'


def _normalize_status(value: str) -> str:
    """标准化处理状态"""
    if not value:
        return '未处理'

    value = value.strip()

    if value in VALID_STATUS:
        return value

    if '已' in value and ('解决' in value or '处理' in value):
        return '已处理'
    elif '处理' in value and '未' not in value:
        return '处理中'
    else:
        return '未处理'


def get_import_template_path() -> str:
    """生成导入模板文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '工作表1'

    # 标题行
    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = '设备报修记录导入模板'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = [
        '序号', '故障教室/书院', '报修时间', '周次', '报修类型',
        '报修人\n（不允许填未知）', '报修人学院', '是否为外聘教师',
        '报修方式', '处理人', '是否更换设备', '设备更换备注', '处理情况',
        '故障发生原因', '处理方式、或未处理原因', '若未处理，处理完毕时间', '最终状态'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 示例数据
    example = [
        1, '行者楼408', '2026-05-30', 14, '中控',
        '张三', '理学院', '否',
        '多媒体报修群', '李四', '否', '', '未处理',
        '中控死机', '', '', ''
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=3, column=col, value=val)

    # 说明行
    ws.cell(row=5, column=1, value='填写说明：')
    ws.cell(row=6, column=1, value='1. 故障教室、报修类型、报修人、处理人为必填项')
    ws.cell(row=7, column=1, value='2. 报修类型可选：中控、电脑、投影仪、音响、麦克风、展台、幕布、网络、软件、其他')
    ws.cell(row=8, column=1, value='3. 处理情况可选：未处理、处理中、已处理')
    ws.cell(row=9, column=1, value='4. 是否为外聘教师、是否更换设备：填"是"或"否"')
    ws.cell(row=10, column=1, value='5. 报修时间格式：YYYY-MM-DD HH:MM:SS')

    # 设置列宽
    col_widths = [8, 15, 20, 8, 12, 15, 18, 12, 15, 12, 12, 12, 25, 30, 20, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 保存到内存缓冲区（不写磁盘）
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
